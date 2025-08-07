from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
import os
import json
from typing import Dict, List
from database import Database
from models import Settings
import csv
import aiohttp
from datetime import datetime, date
import gzip
import io
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = FastAPI(title="VulnAnalizer", version="1.0.0")

# Увеличиваем лимиты для загрузки файлов
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Вспомогательная функция для обработки дат
def parse_date(date_str):
    if not date_str or date_str == '':
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None

# Подключение статических файлов
app.mount("/static", StaticFiles(directory="static"), name="static")

# Настройка шаблонов
templates = Jinja2Templates(directory="templates")

# Инициализация базы данных
db = Database()



@app.on_event("startup")
async def startup():
    # Проверяем соединение с базой данных
    await db.test_connection()

@app.on_event("shutdown")
async def shutdown():
    # Закрытие соединений происходит автоматически
    pass

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/api/settings")
async def get_settings():
    try:
        settings = await db.get_settings()
        return {"success": True, "data": settings}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/settings")
async def update_settings(request: Request):
    try:
        data = await request.json()
        await db.update_settings(data)
        return {"success": True, "message": "Settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/epss/upload")
async def upload_epss(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8').splitlines()
        
        # Ищем строку с заголовками (пропускаем метаданные)
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('cve,') or 'cve' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'cve' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        for row in reader:
            records.append({
                'cve': row['cve'],
                'epss': float(row['epss']),
                'percentile': float(row['percentile']),
                'cvss': float(row.get('cvss', 0)) if row.get('cvss') else None,
                'date': row.get('date') or datetime.utcnow().date()
            })
        await db.insert_epss_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('EPSS upload error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/epss/download")
async def download_epss():
    url = "https://epss.empiricalsecurity.com/epss_scores-current.csv.gz"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download: {resp.status}")
                gz_content = await resp.read()
        with gzip.GzipFile(fileobj=io.BytesIO(gz_content)) as gz:
            decoded = gz.read().decode('utf-8').splitlines()
        
        # Ищем строку с заголовками (пропускаем метаданные)
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('cve,') or 'cve' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'cve' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        for row in reader:
            records.append({
                'cve': row['cve'],
                'epss': float(row['epss']),
                'percentile': float(row['percentile']),
                'cvss': float(row.get('cvss', 0)) if row.get('cvss') else None,
                'date': row.get('date') or datetime.utcnow().date()
            })
        await db.insert_epss_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('EPSS download error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/epss/status")
async def epss_status():
    try:
        count = await db.count_epss_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('EPSS status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/health")
async def health_check():
    try:
        await db.test_connection()
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "database": str(e)}

@app.post("/api/exploitdb/upload")
async def upload_exploitdb(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8').splitlines()
        
        # Ищем строку с заголовками
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('id,') or 'id' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'id' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        skipped_count = 0
        for row_num, row in enumerate(reader, start=header_line + 1):
            try:
                records.append({
                    'exploit_id': int(row['id']),
                    'file_path': row.get('file'),
                    'description': row.get('description'),
                    'date_published': parse_date(row.get('date_published')),
                    'author': row.get('author'),
                    'type': row.get('type'),
                    'platform': row.get('platform'),
                    'port': int(row['port']) if row.get('port') and row['port'].isdigit() else None,
                    'date_added': parse_date(row.get('date_added')),
                    'date_updated': parse_date(row.get('date_updated')),
                    'verified': row.get('verified', '0') == '1',
                    'codes': row.get('codes'),
                    'tags': row.get('tags'),
                    'aliases': row.get('aliases'),
                    'screenshot_url': row.get('screenshot_url'),
                    'application_url': row.get('application_url'),
                    'source_url': row.get('source_url')
                })
            except (ValueError, KeyError) as e:
                skipped_count += 1
                print(f"Skipping invalid row {row_num}: {e}, row data: {row}")
                continue
        
        print(f"Total rows processed: {len(decoded) - header_line}")
        print(f"Valid records: {len(records)}")
        print(f"Skipped records: {skipped_count}")
        
        await db.insert_exploitdb_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('ExploitDB upload error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/exploitdb/download")
async def download_exploitdb():
    url = "https://gitlab.com/exploit-database/exploitdb/-/raw/main/files_exploits.csv"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download: {resp.status}")
                content = await resp.text()
        decoded = content.splitlines()
        
        # Ищем строку с заголовками
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('id,') or 'id' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'id' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        skipped_count = 0
        for row_num, row in enumerate(reader, start=header_line + 1):
            try:
                records.append({
                    'exploit_id': int(row['id']),
                    'file_path': row.get('file'),
                    'description': row.get('description'),
                    'date_published': parse_date(row.get('date_published')),
                    'author': row.get('author'),
                    'type': row.get('type'),
                    'platform': row.get('platform'),
                    'port': int(row['port']) if row.get('port') and row['port'].isdigit() else None,
                    'date_added': parse_date(row.get('date_added')),
                    'date_updated': parse_date(row.get('date_updated')),
                    'verified': row.get('verified', '0') == '1',
                    'codes': row.get('codes'),
                    'tags': row.get('tags'),
                    'aliases': row.get('aliases'),
                    'screenshot_url': row.get('screenshot_url'),
                    'application_url': row.get('application_url'),
                    'source_url': row.get('source_url')
                })
            except (ValueError, KeyError) as e:
                skipped_count += 1
                print(f"Skipping invalid row {row_num}: {e}, row data: {row}")
                continue
        
        print(f"Total rows processed: {len(decoded) - header_line}")
        print(f"Valid records: {len(records)}")
        print(f"Skipped records: {skipped_count}")
        
        await db.insert_exploitdb_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('ExploitDB download error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/exploitdb/status")
async def exploitdb_status():
    try:
        count = await db.count_exploitdb_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('ExploitDB status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

def calculate_impact(settings: dict) -> float:
    """Рассчитать Impact на основе настроек"""
    # Веса для критичности ресурса
    resource_weights = {
        'Critical': 0.33,
        'High': 0.25,
        'Medium': 0.15,
        'None': 0.1
    }
    
    # Веса для конфиденциальных данных
    data_weights = {
        'Есть': 0.33,
        'Отсутствуют': 0.1
    }
    
    # Веса для доступа к интернету
    internet_weights = {
        'Доступен': 0.33,
        'Недоступен': 0.1
    }
    
    # Получаем значения из настроек
    resource_criticality = settings.get('impact_resource_criticality', 'Medium')
    confidential_data = settings.get('impact_confidential_data', 'Отсутствуют')
    internet_access = settings.get('impact_internet_access', 'Недоступен')
    
    # Рассчитываем Impact
    impact = (
        resource_weights.get(resource_criticality, 0.15) +
        data_weights.get(confidential_data, 0.1) +
        internet_weights.get(internet_access, 0.1)
    )
    
    return impact

def calculate_risk_score(epss: float, cvss: float, settings: dict) -> dict:
    """Рассчитать риск по формуле: raw_risk = EPSS * (CVSS / 10) * Impact"""
    if epss is None or cvss is None:
        return {
            'raw_risk': None,
            'risk_score': None,
            'calculation_possible': False,
            'impact': None
        }
    
    # Рассчитываем Impact на основе настроек
    impact = calculate_impact(settings)
    
    raw_risk = epss * (cvss / 10) * impact
    risk_score = min(1, raw_risk) * 100
    
    return {
        'raw_risk': raw_risk,
        'risk_score': risk_score,
        'calculation_possible': True,
        'impact': impact
    }



@app.post("/api/hosts/upload")
async def upload_hosts(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8-sig').splitlines()  # Используем utf-8-sig для удаления BOM
        
        # Парсим CSV с разделителем ;
        reader = csv.DictReader(decoded, delimiter=';')
        
        records = []
        for row in reader:
            try:
                # Парсим hostname и IP из поля Host
                host_info = row['Host'].strip('"')
                hostname = host_info.split(' (')[0] if ' (' in host_info else host_info
                ip_address = host_info.split('(')[1].split(')')[0] if ' (' in host_info else ''
                
                # Парсим CVSS (заменяем запятую на точку)
                cvss_str = row['CVSS'].strip('"').replace(',', '.')
                cvss = float(cvss_str) if cvss_str and cvss_str != '' else None
                
                records.append({
                    'hostname': hostname,
                    'ip_address': ip_address,
                    'cve': row['CVE'].strip('"'),
                    'cvss': cvss,
                    'criticality': row['Criticality'].strip('"'),
                    'status': row['status'].strip('"')
                })
            except Exception as row_error:
                print(f"Error processing row: {row}, error: {row_error}")
                continue
        
        await db.insert_hosts_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('Hosts upload error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/hosts/status")
async def hosts_status():
    try:
        count = await db.count_hosts_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('Hosts status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/hosts/search")
async def search_hosts(
    hostname: str = None,
    cve: str = None,
    ip_address: str = None,
    criticality: str = None,
    exploits_only: bool = False
):
    try:
        results = await db.search_hosts(hostname, cve, ip_address, criticality)
        
        # Фильтруем по эксплойтам если нужно
        if exploits_only:
            filtered_results = []
            for host in results:
                if host.get('has_exploits', False):
                    filtered_results.append(host)
            results = filtered_results
        
        return {"success": True, "data": results}
    except Exception as e:
        print('Hosts search error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/hosts/update-data")
async def update_hosts_data():
    """Обновить данные EPSS и эксплойтов для всех хостов"""
    try:
        updated_count = await db.update_hosts_epss_and_exploits()
        return {
            "success": True, 
            "message": f"Обновлено {updated_count} записей хостов",
            "updated_count": updated_count
        }
    except Exception as e:
        print('Hosts data update error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/hosts/{host_id}/risk")
async def calculate_host_risk(host_id: int):
    try:
        # Получаем данные хоста с расширенной информацией
        host_data = await db.get_host_by_id(host_id)
        if not host_data:
            return {
                "success": False,
                "error": "Host not found",
                "host_id": host_id
            }
        
        # Получаем данные EPSS для CVE хоста (для дополнительной информации)
        epss_data = await db.get_epss_by_cve(host_data['cve'])
        
        # Получаем данные ExploitDB для CVE хоста (для дополнительной информации)
        exploitdb_data = await db.get_exploitdb_by_cve(host_data['cve'])
        
        # Формируем данные о риске из сохраненных данных хоста
        risk_data = None
        if host_data.get('risk_score') is not None:
            risk_data = {
                'raw_risk': host_data.get('risk_raw'),
                'risk_score': host_data.get('risk_score'),
                'calculation_possible': True,
                'impact': host_data.get('impact_score')
            }
        
        return {
            "success": True,
            "host": host_data,
            "epss": epss_data,
            "exploitdb": exploitdb_data,
            "risk": risk_data
        }
    except Exception as e:
        print(f'Host risk calculation error for host {host_id}:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e),
            "host_id": host_id
        }

def create_excel_file(hosts_data: List[Dict]) -> BytesIO:
    """Создает Excel файл с данными хостов"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hosts Data"
    
    # Заголовки с новыми полями
    headers = [
        "ID", "Hostname", "IP Address", "CVE", "CVSS", 
        "Criticality", "Status", "EPSS Score", "EPSS Percentile", 
        "Risk Score", "Risk Raw", "Impact Score",
        "Exploits Count", "Verified Exploits", "Has Exploits", 
        "Last Exploit Date", "EPSS Updated", "Exploits Updated", "Risk Updated"
    ]
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Добавляем данные
    for row, host in enumerate(hosts_data, 2):
        ws.cell(row=row, column=1, value=host.get('id'))
        ws.cell(row=row, column=2, value=host.get('hostname'))
        ws.cell(row=row, column=3, value=host.get('ip_address'))
        ws.cell(row=row, column=4, value=host.get('cve'))
        ws.cell(row=row, column=5, value=host.get('cvss'))
        ws.cell(row=row, column=6, value=host.get('criticality'))
        ws.cell(row=row, column=7, value=host.get('status'))
        
        # EPSS Score с форматированием в процентах
        epss_score = host.get('epss_score')
        if epss_score is not None:
            ws.cell(row=row, column=8, value=epss_score).number_format = '0.00%'
        else:
            ws.cell(row=row, column=8, value='N/A')
        
        # EPSS Percentile с форматированием в процентах
        epss_percentile = host.get('epss_percentile')
        if epss_percentile is not None:
            ws.cell(row=row, column=9, value=epss_percentile).number_format = '0.00%'
        else:
            ws.cell(row=row, column=9, value='N/A')
        
        # Risk Score с форматированием в процентах
        risk_score = host.get('risk_score')
        if risk_score is not None:
            ws.cell(row=row, column=10, value=risk_score).number_format = '0.00%'
        else:
            ws.cell(row=row, column=10, value='N/A')
        
        # Risk Raw
        risk_raw = host.get('risk_raw')
        if risk_raw is not None:
            ws.cell(row=row, column=11, value=risk_raw).number_format = '0.0000'
        else:
            ws.cell(row=row, column=11, value='N/A')
        
        # Impact Score
        impact_score = host.get('impact_score')
        if impact_score is not None:
            ws.cell(row=row, column=12, value=impact_score).number_format = '0.0000'
        else:
            ws.cell(row=row, column=12, value='N/A')
        
        # Exploits Count
        exploits_count = host.get('exploits_count', 0)
        ws.cell(row=row, column=13, value=exploits_count)
        
        # Verified Exploits Count
        verified_exploits = host.get('verified_exploits_count', 0)
        ws.cell(row=row, column=14, value=verified_exploits)
        
        # Has Exploits
        has_exploits = host.get('has_exploits', False)
        ws.cell(row=row, column=15, value='Да' if has_exploits else 'Нет')
        
        # Last Exploit Date
        last_exploit_date = host.get('last_exploit_date')
        if last_exploit_date:
            ws.cell(row=row, column=16, value=last_exploit_date)
        else:
            ws.cell(row=row, column=16, value='N/A')
        
        # EPSS Updated
        epss_updated = host.get('epss_updated_at')
        if epss_updated:
            ws.cell(row=row, column=17, value=epss_updated)
        else:
            ws.cell(row=row, column=17, value='N/A')
        
        # Exploits Updated
        exploits_updated = host.get('exploits_updated_at')
        if exploits_updated:
            ws.cell(row=row, column=18, value=exploits_updated)
        else:
            ws.cell(row=row, column=18, value='N/A')
        
        # Risk Updated
        risk_updated = host.get('risk_updated_at')
        if risk_updated:
            ws.cell(row=row, column=19, value=risk_updated)
        else:
            ws.cell(row=row, column=19, value='N/A')
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.get("/api/hosts/export")
async def export_hosts(
    hostname: str = None,
    cve: str = None,
    ip_address: str = None,
    criticality: str = None,
    exploits_only: bool = False
):
    """Экспорт хостов в Excel файл"""
    try:
        # Получаем данные хостов с теми же фильтрами, что и в поиске
        hosts = await db.search_hosts(hostname, cve, ip_address, criticality)
        
        # Фильтруем по эксплойтам если нужно
        if exploits_only:
            filtered_hosts = []
            for host in hosts:
                if host.get('has_exploits', False):
                    filtered_hosts.append(host)
            hosts = filtered_hosts
        
        if not hosts:
            return {
                "success": False,
                "error": "No hosts found for export"
            }
        
        # Создаем Excel файл с уже рассчитанными данными
        excel_file = create_excel_file(hosts)
        
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hosts_export_{timestamp}.xlsx"
        
        # Возвращаем файл напрямую
        from fastapi.responses import Response
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print('Hosts export error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/hosts/clear")
async def clear_hosts():
    """Очистка таблицы хостов"""
    try:
        await db.clear_hosts()
        return {"success": True, "message": "Hosts table cleared successfully"}
    except Exception as e:
        print('Clear hosts error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/epss/clear")
async def clear_epss():
    """Очистка таблицы EPSS"""
    try:
        await db.clear_epss()
        return {"success": True, "message": "EPSS table cleared successfully"}
    except Exception as e:
        print('Clear EPSS error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/exploitdb/clear")
async def clear_exploitdb():
    """Очистка таблицы ExploitDB"""
    try:
        await db.clear_exploitdb()
        return {"success": True, "message": "ExploitDB table cleared successfully"}
    except Exception as e:
        print('Clear ExploitDB error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        } 