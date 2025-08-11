from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Depends, Form
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import os
import json
from typing import Dict, List, Optional
from database import Database
from models import Settings
from vm_integration import VMMaxPatrolIntegration
import csv
import aiohttp
from datetime import datetime, date, timedelta
import gzip
import io
import traceback
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import jwt
from pydantic import BaseModel
import ipaddress
import tempfile
import shutil

app = FastAPI(title="VulnAnalizer", version="1.0.0")

# Настройки JWT
SECRET_KEY = os.getenv('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Безопасность
security = HTTPBearer()

# Увеличиваем лимиты для загрузки файлов
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Вспомогательная функция для обработки дат
def parse_date(date_str):
    if not date_str or date_str == '':
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return None

def is_valid_ip(ip_str):
    """Проверить, является ли строка валидным IP адресом"""
    if not ip_str or ip_str.strip() == '':
        return False
    try:
        ipaddress.ip_address(ip_str.strip())
        return True
    except ValueError:
        return False

def split_csv_automatically(content: str, max_lines: int = 1000000, delimiter: str = ';') -> list:
    """
    Автоматически разделить CSV контент на части
    
    Args:
        content (str): Содержимое CSV файла
        max_lines (int): Максимальное количество строк в части
        delimiter (str): Разделитель CSV
    
    Returns:
        list: Список частей CSV (каждая часть - строка с содержимым)
    """
    lines = content.splitlines()
    total_lines = len(lines)
    
    if total_lines <= max_lines:
        return [content]  # Файл не нужно разделять
    
    # Читаем заголовки
    headers = lines[0]
    data_lines = lines[1:]
    
    parts = []
    for i in range(0, len(data_lines), max_lines):
        part_lines = data_lines[i:i + max_lines]
        part_content = headers + '\n' + '\n'.join(part_lines)
        parts.append(part_content)
    
    return parts

def split_file_by_size(content: str, max_size_mb: int = 100) -> list:
    """Разделить файл на части по размеру в мегабайтах"""
    max_size_bytes = max_size_mb * 1024 * 1024
    parts = []
    
    if len(content) <= max_size_bytes:
        return [content]
    
    # Находим заголовки (первая строка)
    lines = content.splitlines()
    if not lines:
        return [content]
    
    header = lines[0]
    data_lines = lines[1:]
    
    current_part = [header]
    current_size = len(header.encode('utf-8'))
    
    for line in data_lines:
        line_size = len(line.encode('utf-8'))
        
        if current_size + line_size > max_size_bytes and current_part:
            # Завершаем текущую часть
            parts.append('\n'.join(current_part))
            current_part = [header]  # Начинаем новую часть с заголовка
            current_size = len(header.encode('utf-8'))
        
        current_part.append(line)
        current_size += line_size
    
    # Добавляем последнюю часть
    if current_part:
        parts.append('\n'.join(current_part))
    
    return parts

def update_import_progress(status, current_step, total_steps=None, current_step_progress=None, 
                          total_records=None, processed_records=None, error_message=None,
                          total_parts=None, current_part=None, total_files_processed=None, current_file_records=None):
    """Обновить прогресс импорта"""
    global import_progress
    import_progress.update({
        'status': status,
        'current_step': current_step,
        'error_message': error_message
    })
    
    if total_steps is not None:
        import_progress['total_steps'] = total_steps
    if current_step_progress is not None:
        import_progress['current_step_progress'] = current_step_progress
    if total_records is not None:
        import_progress['total_records'] = total_records
    if processed_records is not None:
        import_progress['processed_records'] = processed_records
    if total_parts is not None:
        import_progress['total_parts'] = total_parts
    if current_part is not None:
        import_progress['current_part'] = current_part
    if total_files_processed is not None:
        import_progress['total_files_processed'] = total_files_processed
    if current_file_records is not None:
        import_progress['current_file_records'] = current_file_records
    
    if import_progress['start_time'] is None and status != 'idle':
        import_progress['start_time'] = datetime.now()
    
    # Рассчитываем правильный процент прогресса
    overall_progress = 0
    if import_progress['total_records'] > 0:
        overall_progress = min(100, (import_progress['processed_records'] / import_progress['total_records']) * 100)
    
    # Формируем информацию о текущем файле
    current_file_info = ""
    if import_progress['current_part'] and import_progress['total_parts']:
        current_file_info = f"Файл {import_progress['current_part']} из {import_progress['total_parts']}"
    
    # Формируем детальное описание текущего шага
    detailed_step = current_step
    if current_file_info:
        detailed_step = f"{current_file_info}: {current_step}"
    
    print(f"📊 Import Progress: {status} - {detailed_step} - {overall_progress:.1f}%")

def extract_compressed_file(file_content: bytes, filename: str) -> str:
    """Извлечь содержимое из сжатого файла"""
    try:
        # Определяем тип архива по расширению
        file_ext = filename.lower()
        
        if file_ext.endswith('.zip'):
            # Обработка ZIP файла
            with zipfile.ZipFile(io.BytesIO(file_content), 'r') as zip_file:
                # Ищем CSV файл в архиве
                csv_files = [f for f in zip_file.namelist() if f.lower().endswith('.csv')]
                if not csv_files:
                    raise Exception("CSV файл не найден в ZIP архиве")
                
                # Берем первый CSV файл
                csv_filename = csv_files[0]
                with zip_file.open(csv_filename) as csv_file:
                    content = csv_file.read()
                    return content.decode('utf-8-sig')
        
        elif file_ext.endswith('.gz') or file_ext.endswith('.gzip'):
            # Обработка GZIP файла
            with gzip.GzipFile(fileobj=io.BytesIO(file_content)) as gz_file:
                content = gz_file.read()
                return content.decode('utf-8-sig')
        

        
        else:
            # Обычный CSV файл
            return file_content.decode('utf-8-sig')
    
    except Exception as e:
        raise Exception(f"Ошибка при извлечении файла: {str(e)}")

def estimate_remaining_time(start_time, processed_records, total_records):
    """Оценить оставшееся время"""
    if processed_records == 0:
        return None
    
    elapsed_time = (datetime.now() - start_time).total_seconds()
    records_per_second = processed_records / elapsed_time
    remaining_records = total_records - processed_records
    
    if records_per_second > 0:
        remaining_seconds = remaining_records / records_per_second
        return remaining_seconds
    
    return None

# Подключение статических файлов
app.mount("/static", StaticFiles(directory="static"), name="static")

# Настройка шаблонов
templates = Jinja2Templates(directory="templates")

# Инициализация базы данных
db = Database()

# Глобальные переменные для отслеживания прогресса импорта
import_progress = {
    'status': 'idle',  # idle, uploading, extracting, splitting, processing, inserting, completed, error
    'current_step': '',
    'progress': 0,
    'total_steps': 0,
    'current_step_progress': 0,
    'total_records': 0,
    'processed_records': 0,
    'error_message': None,
    'start_time': None,
    'estimated_time': None,
    'total_parts': 0,
    'current_part': 0,
    'total_files_processed': 0,
    'current_file_records': 0
}

# JWT функции
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """Создание JWT токена"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str) -> Optional[dict]:
    """Проверка JWT токена"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.PyJWTError:
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Получение текущего пользователя из токена"""
    token = credentials.credentials
    payload = verify_token(token)
    if payload is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    username = payload.get("sub")
    if username is None:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.get_user_by_username(username)
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

# Pydantic модели для пользователей
class UserCreate(BaseModel):
    username: str
    email: Optional[str] = None
    password: str
    is_admin: bool = False

class UserUpdate(BaseModel):
    username: str
    email: Optional[str] = None
    is_active: bool
    is_admin: bool

class PasswordUpdate(BaseModel):
    password: str


@app.on_event("startup")
async def startup():
    # Проверяем соединение с базой данных
    await db.test_connection()
    # Инициализируем админа при первом запуске
    await db.initialize_admin_user()

@app.on_event("shutdown")
async def shutdown():
    # Закрытие соединений происходит автоматически
    pass

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Страница входа"""
    return templates.TemplateResponse("login.html", {"request": request})

@app.get("/api/settings")
async def get_settings():
    try:
        settings = await db.get_settings()
        return {"success": True, "data": settings}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/settings")
async def update_settings(request: Request):
    try:
        data = await request.json()
        await db.update_settings(data)
        return {"success": True, "message": "Settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/epss/upload")
async def upload_epss(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8').splitlines()
        
        # Ищем строку с заголовками (пропускаем метаданные)
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('cve,') or 'cve' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'cve' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        for row in reader:
            records.append({
                'cve': row['cve'],
                'epss': float(row['epss']),
                'percentile': float(row['percentile']),
                'cvss': float(row.get('cvss', 0)) if row.get('cvss') else None,
                'date': row.get('date') or datetime.utcnow().date()
            })
        await db.insert_epss_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('EPSS upload error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/epss/download")
async def download_epss():
    url = "https://epss.empiricalsecurity.com/epss_scores-current.csv.gz"
    try:
        print("🔄 Starting EPSS download...")
        
        # Увеличиваем таймауты для больших файлов
        timeout = aiohttp.ClientTimeout(total=300, connect=60)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            print(f"📥 Downloading from {url}")
            async with session.get(url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download: {resp.status} - {resp.reason}")
                
                print("📦 Reading compressed content...")
                gz_content = await resp.read()
                print(f"📊 Downloaded {len(gz_content)} bytes")
        
        print("🔓 Decompressing content...")
        with gzip.GzipFile(fileobj=io.BytesIO(gz_content)) as gz:
            decoded = gz.read().decode('utf-8').splitlines()
        
        print(f"📄 Decompressed {len(decoded)} lines")
        
        # Ищем строку с заголовками (пропускаем метаданные)
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('cve,') or 'cve' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'cve' column")
        
        print(f"📋 Found header at line {header_line}")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        print("🔄 Processing CSV records...")
        records = []
        processed_count = 0
        
        for row in reader:
            try:
                records.append({
                    'cve': row['cve'],
                    'epss': float(row['epss']),
                    'percentile': float(row['percentile']),
                    'cvss': float(row.get('cvss', 0)) if row.get('cvss') else None,
                    'date': row.get('date') or datetime.utcnow().date()
                })
                processed_count += 1
                
                # Показываем прогресс каждые 10000 записей
                if processed_count % 10000 == 0:
                    print(f"📊 Processed {processed_count} records...")
                    
            except (ValueError, KeyError) as e:
                print(f"⚠️ Skipping invalid row: {e}, row data: {row}")
                continue
        
        print(f"✅ Processed {len(records)} valid records")
        print("💾 Inserting records into database...")
        
        await db.insert_epss_records(records)
        
        print("🎉 EPSS download and processing completed successfully")
        return {"success": True, "count": len(records)}
        
    except Exception as e:
        error_msg = f"EPSS download error: {str(e)}"
        print(error_msg)
        print('Full traceback:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=error_msg)

@app.get("/api/epss/status")
async def epss_status():
    try:
        count = await db.count_epss_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('EPSS status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/health")
async def health_check():
    """Проверка состояния сервиса"""
    try:
        # Проверяем соединение с базой данных
        await db.test_connection()
        
        # Проверяем, не занят ли сервис обработкой большого файла
        global import_progress
        if import_progress['status'] in ['uploading', 'extracting', 'processing', 'inserting']:
            return {
                "status": "busy", 
                "database": "connected",
                "current_operation": import_progress['current_step'],
                "progress": import_progress['progress']
            }
        
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "database": str(e)}

@app.get("/api/version")
async def get_version():
    """Get application version"""
    try:
        with open('/app/VERSION', 'r') as f:
            version = f.read().strip()
        return {"version": version}
    except Exception as e:
        return {"version": "unknown"}

@app.get("/api/system/status")
async def get_system_status():
    """Get system status and memory usage"""
    import psutil
    import os
    
    try:
        # Получаем информацию о памяти
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        
        # Получаем информацию о процессе
        process = psutil.Process(os.getpid())
        
        return {
            "memory": {
                "total_mb": memory.total // (1024 * 1024),
                "available_mb": memory.available // (1024 * 1024),
                "used_mb": memory.used // (1024 * 1024),
                "percent": memory.percent
            },
            "disk": {
                "total_gb": disk.total // (1024 * 1024 * 1024),
                "free_gb": disk.free // (1024 * 1024 * 1024),
                "used_gb": disk.used // (1024 * 1024 * 1024),
                "percent": (disk.used / disk.total) * 100
            },
            "process": {
                "memory_mb": process.memory_info().rss // (1024 * 1024),
                "cpu_percent": process.cpu_percent(),
                "threads": process.num_threads()
            }
        }
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/exploitdb/upload")
async def upload_exploitdb(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8').splitlines()
        
        # Ищем строку с заголовками
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('id,') or 'id' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'id' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        skipped_count = 0
        for row_num, row in enumerate(reader, start=header_line + 1):
            try:
                records.append({
                    'exploit_id': int(row['id']),
                    'file_path': row.get('file'),
                    'description': row.get('description'),
                    'date_published': parse_date(row.get('date_published')),
                    'author': row.get('author'),
                    'type': row.get('type'),
                    'platform': row.get('platform'),
                    'port': int(row['port']) if row.get('port') and row['port'].isdigit() else None,
                    'date_added': parse_date(row.get('date_added')),
                    'date_updated': parse_date(row.get('date_updated')),
                    'verified': row.get('verified', '0') == '1',
                    'codes': row.get('codes'),
                    'tags': row.get('tags'),
                    'aliases': row.get('aliases'),
                    'screenshot_url': row.get('screenshot_url'),
                    'application_url': row.get('application_url'),
                    'source_url': row.get('source_url')
                })
            except (ValueError, KeyError) as e:
                skipped_count += 1
                print(f"Skipping invalid row {row_num}: {e}, row data: {row}")
                continue
        
        print(f"Total rows processed: {len(decoded) - header_line}")
        print(f"Valid records: {len(records)}")
        print(f"Skipped records: {skipped_count}")
        
        await db.insert_exploitdb_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('ExploitDB upload error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/exploitdb/download")
async def download_exploitdb():
    url = "https://gitlab.com/exploit-database/exploitdb/-/raw/main/files_exploits.csv"
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    raise Exception(f"Failed to download: {resp.status}")
                content = await resp.text()
        decoded = content.splitlines()
        
        # Ищем строку с заголовками
        header_line = None
        for i, line in enumerate(decoded):
            if line.startswith('id,') or 'id' in line.split(',')[0]:
                header_line = i
                break
        
        if header_line is None:
            raise Exception("Could not find header line with 'id' column")
        
        # Создаем CSV reader начиная с найденной строки заголовков
        reader = csv.DictReader(decoded[header_line:])
        
        records = []
        skipped_count = 0
        for row_num, row in enumerate(reader, start=header_line + 1):
            try:
                records.append({
                    'exploit_id': int(row['id']),
                    'file_path': row.get('file'),
                    'description': row.get('description'),
                    'date_published': parse_date(row.get('date_published')),
                    'author': row.get('author'),
                    'type': row.get('type'),
                    'platform': row.get('platform'),
                    'port': int(row['port']) if row.get('port') and row['port'].isdigit() else None,
                    'date_added': parse_date(row.get('date_added')),
                    'date_updated': parse_date(row.get('date_updated')),
                    'verified': row.get('verified', '0') == '1',
                    'codes': row.get('codes'),
                    'tags': row.get('tags'),
                    'aliases': row.get('aliases'),
                    'screenshot_url': row.get('screenshot_url'),
                    'application_url': row.get('application_url'),
                    'source_url': row.get('source_url')
                })
            except (ValueError, KeyError) as e:
                skipped_count += 1
                print(f"Skipping invalid row {row_num}: {e}, row data: {row}")
                continue
        
        print(f"Total rows processed: {len(decoded) - header_line}")
        print(f"Valid records: {len(records)}")
        print(f"Skipped records: {skipped_count}")
        
        await db.insert_exploitdb_records(records)
        return {"success": True, "count": len(records)}
    except Exception as e:
        print('ExploitDB download error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/exploitdb/status")
async def exploitdb_status():
    try:
        count = await db.count_exploitdb_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('ExploitDB status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

def calculate_impact(settings: dict) -> float:
    """Рассчитать Impact на основе настроек"""
    # Веса для критичности ресурса
    resource_weights = {
        'Critical': 0.33,
        'High': 0.25,
        'Medium': 0.15,
        'None': 0.1
    }
    
    # Веса для конфиденциальных данных
    data_weights = {
        'Есть': 0.33,
        'Отсутствуют': 0.1
    }
    
    # Веса для доступа к интернету
    internet_weights = {
        'Доступен': 0.33,
        'Недоступен': 0.1
    }
    
    # Получаем значения из настроек
    resource_criticality = settings.get('impact_resource_criticality', 'Medium')
    confidential_data = settings.get('impact_confidential_data', 'Отсутствуют')
    internet_access = settings.get('impact_internet_access', 'Недоступен')
    
    # Рассчитываем Impact
    impact = (
        resource_weights.get(resource_criticality, 0.15) +
        data_weights.get(confidential_data, 0.1) +
        internet_weights.get(internet_access, 0.1)
    )
    
    return impact

def calculate_risk_score(epss: float, cvss: float, settings: dict) -> dict:
    """Рассчитать риск по формуле: raw_risk = EPSS * (CVSS / 10) * Impact"""
    if epss is None or cvss is None:
        return {
            'raw_risk': None,
            'risk_score': None,
            'calculation_possible': False,
            'impact': None
        }
    
    # Рассчитываем Impact на основе настроек
    impact = calculate_impact(settings)
    
    raw_risk = epss * (cvss / 10) * impact
    risk_score = min(1, raw_risk) * 100
    
    return {
        'raw_risk': raw_risk,
        'risk_score': risk_score,
        'calculation_possible': True,
        'impact': impact
    }



@app.post("/api/hosts/upload")
async def upload_hosts(file: UploadFile = File(...)):
    """Загрузить и импортировать хосты из файла с автоматическим разделением больших файлов"""
    global import_progress
    
    try:
        # Сбрасываем прогресс
        update_import_progress('uploading', 'Загрузка файла...', total_parts=0, current_part=0)
        
        print(f"🔄 Начинаем импорт файла: {file.filename} ({file.size} байт)")
        
        # Проверяем размер файла (максимум 1GB для стабильности)
        if file.size and file.size > 1024 * 1024 * 1024:  # 1GB
            error_msg = "Файл слишком большой. Максимальный размер: 1GB."
            update_import_progress('error', error_msg, error_message=error_msg)
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Шаг 1: Загрузка файла
        update_import_progress('uploading', 'Загрузка файла...')
        try:
            content = await file.read()
            file_size_mb = len(content) / (1024 * 1024)
            print(f"📦 Файл загружен: {file_size_mb:.2f} МБ")
        except Exception as read_error:
            error_msg = f"Ошибка при чтении файла: {str(read_error)}"
            update_import_progress('error', error_msg, error_message=error_msg)
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Определяем, является ли файл архивом
        is_archive = file.filename.lower().endswith(('.zip', '.gz', '.gzip'))
        
        if is_archive:
            # Шаг 2: Распаковка архива
            update_import_progress('extracting', 'Распаковка архива...')
            try:
                decoded_content = extract_compressed_file(content, file.filename)
                decoded_size_mb = len(decoded_content.encode('utf-8')) / (1024 * 1024)
                print(f"🔓 Архив распакован: {decoded_size_mb:.2f} МБ")
            except Exception as extract_error:
                error_msg = f"Ошибка при распаковке архива: {str(extract_error)}"
                update_import_progress('error', error_msg, error_message=error_msg)
                raise HTTPException(status_code=400, detail=error_msg)
        else:
            # Если не архив, используем содержимое как есть
            decoded_content = content.decode('utf-8-sig')
            decoded_size_mb = len(decoded_content.encode('utf-8')) / (1024 * 1024)
            print(f"📄 Файл не является архивом: {decoded_size_mb:.2f} МБ")
        
        # Шаг 3: Проверяем размер распакованного файла и разделяем если нужно
        if decoded_size_mb > 100:
            update_import_progress('splitting', f'Файл большой ({decoded_size_mb:.1f} МБ), разделяем на части по 100 МБ...')
            try:
                parts = split_file_by_size(decoded_content, 100)
                total_parts = len(parts)
                print(f"✂️ Файл разделен на {total_parts} частей по 100 МБ")
                update_import_progress('splitting', f'Файл разделен на {total_parts} частей по 100 МБ', total_parts=total_parts)
            except Exception as split_error:
                error_msg = f"Ошибка при разделении файла: {str(split_error)}"
                update_import_progress('error', error_msg, error_message=error_msg)
                raise HTTPException(status_code=400, detail=error_msg)
        else:
            # Файл не нужно разделять
            parts = [decoded_content]
            total_parts = 1
            update_import_progress('processing', 'Файл готов к обработке', total_parts=total_parts)
        
        # Шаг 4: Обработка каждой части
        total_records = 0
        total_processed_lines = 0
        
        for part_index, part_content in enumerate(parts, 1):
            try:
                current_part = part_index
                update_import_progress('processing', f'Обработка файла {current_part} из {total_parts}...', 
                                     current_part=current_part)
                
                print(f"📋 Обрабатываем файл {current_part} из {total_parts}")
                
                # Парсим текущую часть
                part_lines = part_content.splitlines()
                part_total_lines = len(part_lines)
                
                # Парсим CSV с разделителем ;
                reader = csv.DictReader(part_lines, delimiter=';')
                
                part_records = []
                part_processed_lines = 0
                batch_size = 1000
                start_time = datetime.now()
                
                for row in reader:
                    try:
                        # Проверяем время выполнения (максимум 10 минут на файл)
                        if (datetime.now() - start_time).total_seconds() > 600:  # 10 минут
                            error_msg = f"Превышено время обработки файла {current_part} (10 минут). Попробуйте файл меньшего размера."
                            update_import_progress('error', error_msg, error_message=error_msg)
                            raise HTTPException(status_code=408, detail=error_msg)
                        
                        # Парсим hostname и IP из поля @Host
                        host_info = row['@Host'].strip('"')
                        hostname = host_info.split(' (')[0] if ' (' in host_info else host_info
                        ip_address = host_info.split('(')[1].split(')')[0] if ' (' in host_info else ''
                        
                        # Проверяем валидность IP адреса
                        if ip_address and not is_valid_ip(ip_address):
                            print(f"⚠️ Пропускаем запись с невалидным IP: {ip_address}")
                            part_processed_lines += 1
                            continue
                        
                        # Получаем данные из полей
                        cve = row['Host.@Vulners.CVEs'].strip('"')
                        criticality = row['host.UF_Criticality'].strip('"')
                        zone = row['Host.UF_Zone'].strip('"')
                        os_name = row['Host.OsName'].strip('"')
                        status = 'Active'
                        
                        part_records.append({
                            'hostname': hostname,
                            'ip_address': ip_address,
                            'cve': cve,
                            'cvss': None,
                            'criticality': criticality,
                            'status': status,
                            'os_name': os_name,
                            'zone': zone
                        })
                        
                        part_processed_lines += 1
                        
                        # Обновляем прогресс каждые batch_size строк
                        if part_processed_lines % batch_size == 0:
                            update_import_progress('processing', 
                                                 f'Файл {current_part}/{total_parts}: {part_processed_lines:,}/{part_total_lines:,} строк', 
                                                 processed_records=part_processed_lines, total_records=part_total_lines,
                                                 current_file_records=len(part_records))
                            print(f"📊 Часть {current_part}: {part_processed_lines:,}/{part_total_lines:,} строк, {len(part_records):,} записей")
                        
                    except HTTPException:
                        raise
                    except Exception as row_error:
                        print(f"⚠️ Ошибка обработки строки {part_processed_lines} в файле {current_part}: {row_error}")
                        part_processed_lines += 1
                        continue
                
                print(f"✅ Файл {current_part} обработан: {len(part_records):,} записей")
                
                # Шаг 5: Вставка файла в базу данных
                update_import_progress('inserting', f'Сохранение файла {current_part} из {total_parts}...', 
                                     current_file_records=len(part_records))
                
                try:
                    await db.insert_hosts_records_with_progress(part_records, update_import_progress)
                    print(f"💾 Файл {current_part} сохранен в базу данных")
                    
                except Exception as db_error:
                    error_msg = f"Ошибка при сохранении файла {current_part}: {str(db_error)}"
                    update_import_progress('error', error_msg, error_message=error_msg)
                    raise HTTPException(status_code=500, detail=error_msg)
                
                # Обновляем общие счетчики
                total_records += len(part_records)
                total_processed_lines += part_processed_lines
                
                # Обновляем прогресс
                update_import_progress('processing', f'Файл {current_part} из {total_parts} завершен', 
                                     total_records=total_records, processed_records=total_processed_lines, total_files_processed=current_part)
                
            except HTTPException:
                raise
            except Exception as part_error:
                error_msg = f"Ошибка при обработке файла {current_part}: {str(part_error)}"
                update_import_progress('error', error_msg, error_message=error_msg)
                raise HTTPException(status_code=500, detail=error_msg)
        
        # Завершение
        update_import_progress('completed', 'Импорт завершен', total_records=total_records, processed_records=total_processed_lines, 
                              total_files_processed=total_parts)
        print(f"🎉 Импорт успешно завершен: {total_records:,} записей из {total_processed_lines:,} строк в {total_parts} файлах")
        
        return {
            "success": True, 
            "count": total_records, 
            "total_processed": total_processed_lines,
            "total_parts": total_parts,
            "message": f"Файл автоматически разделен на {total_parts} файлов по 100 МБ и успешно импортирован"
        }
        
    except HTTPException:
        # Перебрасываем HTTP исключения как есть
        raise
    except Exception as e:
        error_msg = f"Неожиданная ошибка при импорте: {str(e)}"
        update_import_progress('error', error_msg, error_message=error_msg)
        print(f'❌ Hosts upload error: {traceback.format_exc()}')
        raise HTTPException(status_code=500, detail=error_msg)

@app.get("/api/hosts/status")
async def hosts_status():
    try:
        count = await db.count_hosts_records()
        return {"success": True, "count": count}
    except Exception as e:
        print('Hosts status error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/hosts/import-progress")
async def get_import_progress():
    """Получить текущий прогресс импорта хостов"""
    global import_progress
    
    # Рассчитываем оставшееся время
    estimated_time = None
    if (import_progress['start_time'] and 
        import_progress['processed_records'] > 0 and 
        import_progress['total_records'] > 0):
        estimated_time = estimate_remaining_time(
            import_progress['start_time'],
            import_progress['processed_records'],
            import_progress['total_records']
        )
    
    # Рассчитываем правильный процент прогресса
    overall_progress = 0
    if import_progress['total_records'] > 0:
        overall_progress = min(100, (import_progress['processed_records'] / import_progress['total_records']) * 100)
    
    # Формируем информацию о текущем файле
    current_file_info = ""
    if import_progress['current_part'] and import_progress['total_parts']:
        current_file_info = f"Файл {import_progress['current_part']} из {import_progress['total_parts']}"
    
    # Формируем детальное описание текущего шага
    detailed_step = import_progress['current_step']
    if current_file_info:
        detailed_step = f"{current_file_info}: {import_progress['current_step']}"
    
    return {
        "status": import_progress['status'],
        "current_step": detailed_step,
        "progress": overall_progress,
        "total_steps": import_progress['total_steps'],
        "current_step_progress": import_progress['current_step_progress'],
        "total_records": import_progress['total_records'],
        "processed_records": import_progress['processed_records'],
        "error_message": import_progress['error_message'],
        "estimated_time": estimated_time,
        "total_parts": import_progress['total_parts'],
        "current_part": import_progress['current_part'],
        "total_files_processed": import_progress['total_files_processed'],
        "current_file_records": import_progress['current_file_records'],
        "current_file_info": current_file_info,
        "overall_progress": overall_progress
    }

@app.get("/api/hosts/import-limits")
async def get_import_limits():
    """Получить информацию о лимитах импорта"""
    return {
        "max_file_size_mb": 1024,  # 1GB
        "max_processing_time_minutes": 10,
        "recommended_file_size_mb": 100,
        "auto_split_size_mb": 100,
        "message": "Файлы больше 100 МБ автоматически разделяются на части по 100 МБ"
    }

@app.get("/api/hosts/search")
async def search_hosts(
    hostname: str = None,
    cve: str = None,
    ip_address: str = None,
    criticality: str = None,
    exploits_only: bool = False
):
    try:
        results = await db.search_hosts(hostname, cve, ip_address, criticality)
        
        # Фильтруем по эксплойтам если нужно
        if exploits_only:
            filtered_results = []
            for host in results:
                if host.get('has_exploits', False):
                    filtered_results.append(host)
            results = filtered_results
        
        return {"success": True, "data": results}
    except Exception as e:
        print('Hosts search error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/hosts/update-data")
async def update_hosts_data():
    """Обновить данные EPSS и эксплойтов для всех хостов"""
    try:
        updated_count = await db.update_hosts_epss_and_exploits()
        return {
            "success": True, 
            "message": f"Обновлено {updated_count} записей хостов",
            "updated_count": updated_count
        }
    except Exception as e:
        print('Hosts data update error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/hosts/{host_id}/risk")
async def calculate_host_risk(host_id: int):
    try:
        # Получаем данные хоста с расширенной информацией
        host_data = await db.get_host_by_id(host_id)
        if not host_data:
            return {
                "success": False,
                "error": "Host not found",
                "host_id": host_id
            }
        
        # Получаем данные EPSS для CVE хоста (для дополнительной информации)
        epss_data = await db.get_epss_by_cve(host_data['cve'])
        
        # Получаем данные ExploitDB для CVE хоста (для дополнительной информации)
        exploitdb_data = await db.get_exploitdb_by_cve(host_data['cve'])
        
        # Формируем данные о риске из сохраненных данных хоста
        risk_data = None
        if host_data.get('risk_score') is not None:
            risk_data = {
                'raw_risk': host_data.get('risk_raw'),
                'risk_score': host_data.get('risk_score'),
                'calculation_possible': True,
                'impact': host_data.get('impact_score')
            }
        
        return {
            "success": True,
            "host": host_data,
            "epss": epss_data,
            "exploitdb": exploitdb_data,
            "risk": risk_data
        }
    except Exception as e:
        print(f'Host risk calculation error for host {host_id}:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e),
            "host_id": host_id
        }

def create_excel_file(hosts_data: List[Dict]) -> BytesIO:
    """Создает Excel файл с данными хостов"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hosts Data"
    
    # Заголовки с новыми полями
    headers = [
        "ID", "Hostname", "IP Address", "CVE", "CVSS", 
        "Criticality", "Status", "EPSS Score", "EPSS Percentile", 
        "Risk Score", "Risk Raw", "Impact Score",
        "Exploits Count", "Verified Exploits", "Has Exploits", 
        "Last Exploit Date", "EPSS Updated", "Exploits Updated", "Risk Updated"
    ]
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Добавляем данные
    for row, host in enumerate(hosts_data, 2):
        ws.cell(row=row, column=1, value=host.get('id'))
        ws.cell(row=row, column=2, value=host.get('hostname'))
        ws.cell(row=row, column=3, value=host.get('ip_address'))
        ws.cell(row=row, column=4, value=host.get('cve'))
        ws.cell(row=row, column=5, value=host.get('cvss'))
        ws.cell(row=row, column=6, value=host.get('criticality'))
        ws.cell(row=row, column=7, value=host.get('status'))
        
        # EPSS Score с форматированием в процентах
        epss_score = host.get('epss_score')
        if epss_score is not None:
            ws.cell(row=row, column=8, value=epss_score).number_format = '0.00%'
        else:
            ws.cell(row=row, column=8, value='N/A')
        
        # EPSS Percentile с форматированием в процентах
        epss_percentile = host.get('epss_percentile')
        if epss_percentile is not None:
            ws.cell(row=row, column=9, value=epss_percentile).number_format = '0.00%'
        else:
            ws.cell(row=row, column=9, value='N/A')
        
        # Risk Score с форматированием в процентах
        risk_score = host.get('risk_score')
        if risk_score is not None:
            ws.cell(row=row, column=10, value=risk_score).number_format = '0.00%'
        else:
            ws.cell(row=row, column=10, value='N/A')
        
        # Risk Raw
        risk_raw = host.get('risk_raw')
        if risk_raw is not None:
            ws.cell(row=row, column=11, value=risk_raw).number_format = '0.0000'
        else:
            ws.cell(row=row, column=11, value='N/A')
        
        # Impact Score
        impact_score = host.get('impact_score')
        if impact_score is not None:
            ws.cell(row=row, column=12, value=impact_score).number_format = '0.0000'
        else:
            ws.cell(row=row, column=12, value='N/A')
        
        # Exploits Count
        exploits_count = host.get('exploits_count', 0)
        ws.cell(row=row, column=13, value=exploits_count)
        
        # Verified Exploits Count
        verified_exploits = host.get('verified_exploits_count', 0)
        ws.cell(row=row, column=14, value=verified_exploits)
        
        # Has Exploits
        has_exploits = host.get('has_exploits', False)
        ws.cell(row=row, column=15, value='Да' if has_exploits else 'Нет')
        
        # Last Exploit Date
        last_exploit_date = host.get('last_exploit_date')
        if last_exploit_date:
            ws.cell(row=row, column=16, value=last_exploit_date)
        else:
            ws.cell(row=row, column=16, value='N/A')
        
        # EPSS Updated
        epss_updated = host.get('epss_updated_at')
        if epss_updated:
            ws.cell(row=row, column=17, value=epss_updated)
        else:
            ws.cell(row=row, column=17, value='N/A')
        
        # Exploits Updated
        exploits_updated = host.get('exploits_updated_at')
        if exploits_updated:
            ws.cell(row=row, column=18, value=exploits_updated)
        else:
            ws.cell(row=row, column=18, value='N/A')
        
        # Risk Updated
        risk_updated = host.get('risk_updated_at')
        if risk_updated:
            ws.cell(row=row, column=19, value=risk_updated)
        else:
            ws.cell(row=row, column=19, value='N/A')
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.get("/api/hosts/export")
async def export_hosts(
    hostname: str = None,
    cve: str = None,
    ip_address: str = None,
    criticality: str = None,
    exploits_only: bool = False
):
    """Экспорт хостов в Excel файл"""
    try:
        # Получаем данные хостов с теми же фильтрами, что и в поиске
        hosts = await db.search_hosts(hostname, cve, ip_address, criticality)
        
        # Фильтруем по эксплойтам если нужно
        if exploits_only:
            filtered_hosts = []
            for host in hosts:
                if host.get('has_exploits', False):
                    filtered_hosts.append(host)
            hosts = filtered_hosts
        
        if not hosts:
            return {
                "success": False,
                "error": "No hosts found for export"
            }
        
        # Создаем Excel файл с уже рассчитанными данными
        excel_file = create_excel_file(hosts)
        
        # Формируем имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hosts_export_{timestamp}.xlsx"
        
        # Возвращаем файл напрямую
        from fastapi.responses import Response
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print('Hosts export error:', traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/hosts/clear")
async def clear_hosts():
    """Очистка таблицы хостов"""
    try:
        await db.clear_hosts()
        return {"success": True, "message": "Hosts table cleared successfully"}
    except Exception as e:
        print('Clear hosts error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/epss/clear")
async def clear_epss():
    """Очистка таблицы EPSS"""
    try:
        await db.clear_epss()
        return {"success": True, "message": "EPSS table cleared successfully"}
    except Exception as e:
        print('Clear EPSS error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/exploitdb/clear")
async def clear_exploitdb():
    """Очистка таблицы ExploitDB"""
    try:
        await db.clear_exploitdb()
        return {"success": True, "message": "ExploitDB table cleared successfully"}
    except Exception as e:
        print('Clear ExploitDB error:', traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        } 

# ===== VM MAXPATROL ИНТЕГРАЦИЯ =====

@app.get("/api/vm/settings")
async def get_vm_settings():
    """Получить настройки VM MaxPatrol"""
    try:
        settings = await db.get_vm_settings()
        return {"success": True, "data": settings}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/vm/settings")
async def update_vm_settings(request: Request):
    """Обновить настройки VM MaxPatrol"""
    try:
        data = await request.json()
        await db.update_vm_settings(data)
        return {"success": True, "message": "VM settings updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/vm/test-connection")
async def test_vm_connection(request: Request):
    """Тестировать подключение к VM MaxPatrol"""
    try:
        data = await request.json()
        vm_host = data.get('vm_host', '')
        vm_username = data.get('vm_username', '')
        vm_password = data.get('vm_password', '')
        vm_client_secret = data.get('vm_client_secret', '')
        
        if not all([vm_host, vm_username, vm_password, vm_client_secret]):
            return {
                "success": False,
                "error": "All VM connection parameters are required"
            }
        
        # Создаем экземпляр интеграции
        vm_integration = VMMaxPatrolIntegration(
            host=vm_host,
            username=vm_username,
            password=vm_password,
            client_secret=vm_client_secret
        )
        
        # Тестируем подключение
        result = vm_integration.test_connection()
        
        return {
            "success": True,
            "data": result
        }
        
    except Exception as e:
        print(f"VM connection test error: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/vm/import")
async def import_vm_hosts():
    """Импортировать хосты из VM MaxPatrol"""
    try:
        # Получаем настройки VM
        vm_settings = await db.get_vm_settings()
        
        if vm_settings.get('vm_enabled') != 'true':
            return {
                "success": False,
                "error": "VM integration is not enabled"
            }
        
        vm_host = vm_settings.get('vm_host', '')
        vm_username = vm_settings.get('vm_username', '')
        vm_password = vm_settings.get('vm_password', '')
        vm_client_secret = vm_settings.get('vm_client_secret', '')
        vm_os_filter = vm_settings.get('vm_os_filter', '')
        vm_limit = int(vm_settings.get('vm_limit', '0'))
        
        if not all([vm_host, vm_username, vm_password, vm_client_secret]):
            return {
                "success": False,
                "error": "VM connection parameters are not configured"
            }
        
        # Создаем экземпляр интеграции
        vm_integration = VMMaxPatrolIntegration(
            host=vm_host,
            username=vm_username,
            password=vm_password,
            client_secret=vm_client_secret
        )
        
        # Получаем данные хостов из VM
        hosts_data = vm_integration.get_hosts_data(os_filter=vm_os_filter, limit=vm_limit)
        
        if not hosts_data:
            return {
                "success": False,
                "error": "No hosts data received from VM"
            }
        
        # Импортируем данные в базу
        import_result = await db.import_vm_hosts(hosts_data)
        
        # Обновляем статус импорта
        await db.update_vm_import_status(import_result['total_processed'])
        
        return {
            "success": True,
            "message": f"Successfully imported {import_result['total_processed']} host-CVE combinations",
            "data": import_result
        }
        
    except Exception as e:
        error_msg = str(e)
        print(f"VM import error: {error_msg}")
        
        # Обновляем статус с ошибкой
        await db.update_vm_import_status(0, error_msg)
        
        return {
            "success": False,
            "error": error_msg
        }

@app.get("/api/vm/status")
async def get_vm_status():
    """Получить статус VM интеграции"""
    try:
        vm_settings = await db.get_vm_settings()
        import_status = await db.get_vm_import_status()
        
        return {
            "success": True,
            "data": {
                "settings": vm_settings,
                "import_status": import_status
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) 

# ===== УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ =====

@app.post("/api/users/register")
async def register_user(user: UserCreate):
    """Регистрация нового пользователя"""
    try:
        existing_user = await db.get_user_by_username(user.username)
        if existing_user:
            raise HTTPException(status_code=400, detail="Username already registered")
        
        await db.create_user(user.username, user.password, user.email, user.is_admin)
        return {"success": True, "message": "User registered successfully"}
    except Exception as e:
        print(f"User registration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/users/login")
async def login_user(username: str = Form(...), password: str = Form(...)):
    """Авторизация пользователя"""
    user = await db.get_user_by_username(username)
    if not user or not user.get('password') == password:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer", "user": user}

@app.get("/api/users/me")
async def read_users_me(current_user: dict = Depends(get_current_user)):
    """Получить информацию о текущем пользователе"""
    return current_user

@app.put("/api/users/me", response_model=dict)
async def update_user_me(
    current_user: dict = Depends(get_current_user),
    user_update: UserUpdate = Form(...)
):
    """Обновить информацию о текущем пользователе"""
    try:
        await db.update_user(current_user["id"], user_update.username, user_update.email, user_update.is_active, user_update.is_admin)
        return {"success": True, "message": "User updated successfully"}
    except Exception as e:
        print(f"User update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/users/me/password", response_model=dict)
async def update_user_password_me(
    current_user: dict = Depends(get_current_user),
    password_update: PasswordUpdate = Form(...)
):
    """Обновить пароль текущего пользователя"""
    try:
        await db.update_user_password(current_user["id"], password_update.password)
        return {"success": True, "message": "Password updated successfully"}
    except Exception as e:
        print(f"User password update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/users/all")
async def get_all_users(current_user: dict = Depends(get_current_user)):
    """Получить список всех пользователей (только для админов)"""
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Only administrators can access this endpoint")
    try:
        users = await db.get_all_users()
        return {"success": True, "data": users}
    except Exception as e:
        print(f"Get all users error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/users/{user_id}/update")
async def update_user_by_id(
    user_id: int,
    user_update: UserUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Обновить пользователя по ID (только для админов)"""
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Only administrators can access this endpoint")
    try:
        await db.update_user(user_id, user_update.username, user_update.email, user_update.is_active, user_update.is_admin)
        return {"success": True, "message": "User updated successfully"}
    except Exception as e:
        print(f"Update user by ID error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/users/{user_id}/password")
async def update_user_password_by_id(
    user_id: int,
    password_update: PasswordUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Обновить пароль пользователя по ID (только для админов)"""
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Only administrators can access this endpoint")
    try:
        await db.update_user_password(user_id, password_update.password)
        return {"success": True, "message": "Password updated successfully"}
    except Exception as e:
        print(f"Update user password by ID error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/users/{user_id}/delete")
async def delete_user_by_id(
    user_id: int,
    current_user: dict = Depends(get_current_user)
):
    """Удалить пользователя по ID (только для админов)"""
    if not current_user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Only administrators can access this endpoint")
    try:
        await db.delete_user(user_id)
        return {"success": True, "message": "User deleted successfully"}
    except Exception as e:
        print(f"Delete user by ID error: {e}")
        raise HTTPException(status_code=500, detail=str(e)) 