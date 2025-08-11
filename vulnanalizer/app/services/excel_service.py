"""
Сервис для работы с Excel файлами
"""
from io import BytesIO
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_file(hosts_data: List[Dict]) -> BytesIO:
    """Создает Excel файл с данными хостов"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hosts Data"
    
    # Заголовки с новыми полями
    headers = [
        'ID', 'Hostname', 'IP Address', 'CVE', 'CVSS', 'EPSS', 'Risk Score', 
        'Criticality', 'Status', 'OS Name', 'Zone', 'Has Exploits', 'Impact Score'
    ]
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Стили для границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Записываем данные
    for row, host in enumerate(hosts_data, 2):
        ws.cell(row=row, column=1, value=host.get('id')).border = thin_border
        ws.cell(row=row, column=2, value=host.get('hostname')).border = thin_border
        ws.cell(row=row, column=3, value=host.get('ip_address')).border = thin_border
        ws.cell(row=row, column=4, value=host.get('cve')).border = thin_border
        ws.cell(row=row, column=5, value=host.get('cvss')).border = thin_border
        ws.cell(row=row, column=6, value=host.get('epss')).border = thin_border
        ws.cell(row=row, column=7, value=host.get('risk_score')).border = thin_border
        ws.cell(row=row, column=8, value=host.get('criticality')).border = thin_border
        ws.cell(row=row, column=9, value=host.get('status')).border = thin_border
        ws.cell(row=row, column=10, value=host.get('os_name')).border = thin_border
        ws.cell(row=row, column=11, value=host.get('zone')).border = thin_border
        ws.cell(row=row, column=12, value=host.get('has_exploits')).border = thin_border
        ws.cell(row=row, column=13, value=host.get('impact_score')).border = thin_border
    
    # Автоматическая ширина столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
